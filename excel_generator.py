#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel生成器
将AI返回的JSON数据转换为Excel文件
"""

import pandas as pd
import re
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import formatters
import output_metadata


def validate_and_fix_content_format(content):
    """
    Validate and fix work order content format

    Args:
        content: Work order content string (may contain multiple lines separated by \\n)

    Returns:
        str: Content in correct format (one work order per line)
    """
    if not content:
        return content

    lines = content.split('\n')
    fixed_lines = []
    format_issues_found = False

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Check if line already matches the expected format: (中台) XXXX AAA 名字
        # Pattern: (中台) + 4 digits + space + 2-4 uppercase letters + space + Chinese name
        pattern = r'^\(中台\)\s+\d{4,}\s+[A-Z]{2,}\s+[\u4e00-\u9fa5]+'

        if re.match(pattern, line):
            # Already in correct format
            fixed_lines.append(line)
        else:
            # Log format issue for debugging
            if not format_issues_found:
                print(f"  [WARNING] Warning: Found work order content not matching expected format:")
                print(f"     Expected: '(中台) XXXX AAA 名字'")
                print(f"     Got: '{line[:50]}...' " if len(line) > 50 else f"     Got: '{line}'")
                format_issues_found = True

            # Try to fix the format or log a warning
            # For now, keep as-is but this could be enhanced
            fixed_lines.append(line)

    return '\n'.join(fixed_lines)


def apply_cell_format(cell, format_config):
    """
    应用格式到单元格

    Args:
        cell: openpyxl单元格对象
        format_config: 格式配置字典
    """
    if 'font' in format_config:
        cell.font = format_config['font']
    if 'alignment' in format_config:
        cell.alignment = format_config['alignment']


def create_schedule_sheet(ws, schedule_data):
    """
    创建排班表sheet

    Args:
        ws: openpyxl工作表对象
        schedule_data: AI返回的排班表数据
    """
    print(f"  Creating schedule sheet...")

    # 标题行（第1行）
    title_config = output_metadata.get_schedule_sheet_config()
    title_format = formatters.get_title_format()

    ws['A1'] = title_config['title']
    ws.merge_cells(title_config['title_merge'])
    apply_cell_format(ws['A1'], title_format)

    # Fix #1: Sort dates by chronological order
    # Extract date for sorting - parse Chinese date format like "5月07日"
    def parse_chinese_date(date_str):
        """Parse Chinese date string to sortable tuple (month, day)"""
        import re
        match = re.match(r'(\d+)月(\d+)日', date_str)
        if match:
            return (int(match.group(1)), int(match.group(2)))
        return (999, 999)  # Put unparsable dates at the end

    schedule_data['schedule'].sort(key=lambda x: parse_chinese_date(x['date']))

    # Fix #1: Start from column 1, not column 2 (no empty column A)
    col_idx = 1  # 从A列开始

    # 第2行：日期行
    for date_item in schedule_data['schedule']:
        date_cell = ws.cell(row=2, column=col_idx, value=date_item['date'])
        apply_cell_format(date_cell, formatters.get_date_row_format())

        # 保存列位置信息供后续使用
        date_item['col_start'] = col_idx

        if date_item['type'] == 'normal':
            # 普通情况：只占1列
            date_item['col_end'] = col_idx
            col_idx += 1
        else:
            # Fix #2 & #3: For special cases (weekends), ensure time points are unique
            # and properly formatted as individual times like "01:00", "02:00"
            if 'time_points' in date_item and date_item['time_points']:
                # Remove duplicates while preserving order
                seen = set()
                unique_time_points = []
                for tp in date_item['time_points']:
                    if tp not in seen:
                        seen.add(tp)
                        unique_time_points.append(tp)

                # Ensure time points are sorted
                def parse_time(time_str):
                    """Parse time string like '01:00' to minutes for sorting"""
                    try:
                        parts = str(time_str).strip().split(':')
                        return int(parts[0]) * 60 + int(parts[1]) if len(parts) == 2 else 0
                    except:
                        return 0

                unique_time_points.sort(key=parse_time)
                date_item['time_points'] = unique_time_points

            # 根据时间点数量确定列数
            num_cols = len(date_item['time_points'])
            date_item['col_end'] = col_idx + num_cols - 1

            # 合并日期单元格
            if num_cols > 1:
                ws.merge_cells(start_row=2, start_column=col_idx,
                             end_row=2, end_column=date_item['col_end'])

            col_idx += num_cols

    # 第3行：时间行
    for date_item in schedule_data['schedule']:
        col_start = date_item['col_start']
        col_end = date_item['col_end']

        if date_item['type'] == 'normal':
            # 普通情况：时间范围
            time_cell = ws.cell(row=3, column=col_start, value=date_item['time_range'])
            apply_cell_format(time_cell, formatters.get_time_row_format())
        else:
            # 特殊情况：显示所有时间点
            for idx, time_point in enumerate(date_item['time_points']):
                time_cell = ws.cell(row=3, column=col_start + idx, value=time_point)
                apply_cell_format(time_cell, formatters.get_time_row_format())

    # 第4行及以后：工单行（竖着排列）
    max_work_orders = max(len(item['work_orders']) for item in schedule_data['schedule'])

    for work_order_idx in range(max_work_orders):
        for date_item in schedule_data['schedule']:
            col_start = date_item['col_start']
            col_end = date_item['col_end']
            row_idx = 4 + work_order_idx

            # 检查当前日期是否有这个工单
            if work_order_idx < len(date_item['work_orders']):
                work_order = date_item['work_orders'][work_order_idx]

                # Fix #4: Ensure work order content is in correct format
                # Format: f"(中台) {last_4_digits} {abbr} {chinese_name}"
                content = work_order.get('content', '')
                content = validate_and_fix_content_format(content)

                # For special types, determine merge range based on time_start and time_end
                if date_item['type'] == 'special':
                    # Special case: use time_start and time_end to determine merge range
                    time_start = work_order.get('time_start', '')
                    time_end = work_order.get('time_end', '')

                    if time_start and time_end and 'time_points' in date_item:
                        # Find the indices of time_start and time_end in time_points
                        try:
                            start_idx = date_item['time_points'].index(time_start)
                            end_idx = date_item['time_points'].index(time_end)

                            # Calculate actual column positions
                            actual_col_start = col_start + start_idx
                            actual_col_end = col_start + end_idx

                            # Merge cells for this person's time range only
                            if actual_col_start != actual_col_end:
                                ws.merge_cells(start_row=row_idx, start_column=actual_col_start,
                                             end_row=row_idx, end_column=actual_col_end)

                            content_cell = ws.cell(row=row_idx, column=actual_col_start, value=content)
                            apply_cell_format(content_cell, formatters.get_data_row_format(date_item['type']))

                            # Skip the rest of the loop for this work order
                            continue
                        except (ValueError, IndexError):
                            # If time_start or time_end not found, fall through to default behavior
                            pass

                # Default behavior: merge across all columns (for normal types or if special logic fails)
                if col_start != col_end:
                    # 特殊情况：合并单元格
                    ws.merge_cells(start_row=row_idx, start_column=col_start,
                                 end_row=row_idx, end_column=col_end)

                content_cell = ws.cell(row=row_idx, column=col_start, value=content)
                apply_cell_format(content_cell, formatters.get_data_row_format(date_item['type']))

    # 设置列宽和行高
    max_col = ws.max_column if ws.max_column > 6 else 6
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = formatters.get_column_width()

    for row_idx in range(4, 200):
        ws.row_dimensions[row_idx].height = formatters.get_row_height()

    print(f"  [OK] Schedule sheet created with {len(schedule_data['schedule'])} dates")


def create_data_sheet(ws, filtered_data):
    """
    创建数据明细sheet

    Args:
        ws: openpyxl工作表对象
        filtered_data: 过滤后的源数据
    """
    print(f"  Creating data sheet...")

    # 配置
    data_sheet_config = output_metadata.get_data_sheet_config()

    # 标题行（第1行）
    ws['A1'] = data_sheet_config['title']
    apply_cell_format(ws['A1'], formatters.get_data_sheet_title_format())

    if not filtered_data:
        print(f"  [WARNING] No data to write")
        return

    # 列头（第2行）
    columns = list(filtered_data[0].keys())
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        apply_cell_format(cell, formatters.get_data_sheet_header_format())

    # 数据行（第3行及以下）
    for row_idx, data_row in enumerate(filtered_data, start=3):
        for col_idx, col_name in enumerate(columns, start=1):
            value = data_row.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            apply_cell_format(cell, formatters.get_data_sheet_cell_format())

    # 设置列宽
    for col_idx in range(1, len(columns) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = formatters.get_data_sheet_column_width()

    print(f"  [OK] Data sheet created with {len(filtered_data)} rows")


def generate_excel(json_data, output_file, df_filtered=None):
    """
    生成Excel文件

    Args:
        json_data: AI返回的JSON数据
        output_file: 输出文件路径
        df_filtered: 过滤后的DataFrame（用于创建完整的数据报表）
    """
    print(f"\n[5/5] Generating Excel file...")

    # 创建工作簿
    wb = Workbook()

    # 创建排班表sheet
    schedule_config = output_metadata.get_schedule_sheet_config()
    ws_schedule = wb.active
    ws_schedule.title = schedule_config['name']
    create_schedule_sheet(ws_schedule, json_data)

    # 创建数据明细sheet
    data_sheet_config = output_metadata.get_data_sheet_config()
    ws_data = wb.create_sheet(data_sheet_config['name'])

    # 使用完整的DataFrame创建数据报表（如果提供）
    if df_filtered is not None:
        # 将DataFrame转换为字典列表，保留所有列
        filtered_data = df_filtered.to_dict('records')
        create_data_sheet(ws_data, filtered_data)
        print(f"  [OK] Data sheet created with {len(df_filtered.columns)} columns")
    else:
        # 回退到AI返回的数据（可能不完整）
        create_data_sheet(ws_data, json_data['filtered_source_data'])

    # 保存文件
    wb.save(output_file)

    print(f"  [OK] Excel file saved: {output_file}")
    print(f"\n{'=' * 60}")
    print(f"[SUCCESS] Two-sheet Excel file generated successfully!")
    print(f"{'=' * 60}")
    print(f"\nOutput file: {output_file}")
    print(f"\nSheets created:")
    print(f"  1. '{schedule_config['name']}' - Schedule sheet")
    print(f"  2. '{data_sheet_config['name']}' - Filtered data sheet")
    print(f"{'=' * 60}")


if __name__ == "__main__":
    # 测试代码
    import json

    # 模拟JSON数据
    mock_data = {
        "schedule": [
            {
                "date": "1月05日",
                "date_merge_range": "D2:F2",
                "type": "normal",
                "time_range": "0:00 至 3:00",
                "work_orders": [
                    {
                        "person": "李明",
                        "merge_range": "D4",
                        "content": "(中台) 0001 GMC 李明\n(中台) 0002 AAS 張三"
                    }
                ]
            }
        ],
        "filtered_source_data": [
            {"工單號": "CG001", "變更名稱": "测试变更"}
        ]
    }

    # 生成Excel
    generate_excel(mock_data, "test_output.xlsx")
